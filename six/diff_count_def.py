import os
import cv2
import numpy as np
import openpyxl
import pathlib
from tqdm import tqdm

def diffcountfiles(foldar_name,first_num,last_num,time):

    basename = os.path.basename(foldar_name)
    p_sub = pathlib.Path(foldar_name)
    p_sub_name= str(p_sub.parent)
    A=2-first_num#最上段から表示させるための補正

    for x in ["A","B","C","D","E","F"]:

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'time'
        sheet['B1'] = basename+"_"+x
        wb.save(p_sub_name+'/excel_files/'+basename+'_'+x+'_'+time+'.xlsx')

        for num in tqdm(range (first_num,last_num)):
            file_num= str(num)
            file_name= file_num.zfill(6)+"_diff.jpg"

            img_diff = cv2.imread(p_sub_name+"/crop_imgs/"+basename+'_'+x+'_'+time+"/"+file_name, cv2.IMREAD_GRAYSCALE)
            whitePixels = np.count_nonzero(img_diff)
            sabun = whitePixels

            sabunnum=num+A

            sheet = wb.active
            sheet.cell(row=sabunnum, column=1).value = num
            sheet.cell(row=sabunnum, column=2).value = whitePixels
            wb.save(p_sub_name+'/excel_files/'+basename+'_'+x+'_'+time+'.xlsx')
