import tkinter
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import mkdiffimg_def
import cropimage_def
import diff_count_def
import datetime
import openpyxl
import os

def ask_folder():
    #フォルダ指定
    path = filedialog.askdirectory()
    folder_path.set(path)

def app():
    '''処理実行
    '''
    #時間取得
    now = datetime.datetime.now()
    time = now.strftime("%Y%m%d%H%M")
    logtime =now.strftime("%Y/%m/%d/%H:%M:%S")

    #差分画像作成
    foldar_name = folder_path.get()
    first_num = first.get()
    last_num =last.get()
    th =threshold.get()
    way=radioValue.get()
    Opening = Opening_value.get()

    first_num = first.get()
    last_num =last.get()

    A_1n=A_1.get()
    A_2n=A_2.get()
    A_3n=A_3.get()
    A_4n=A_4.get()
    Acrop=(A_1n,A_2n,A_3n,A_4n)

    B_1n=B_1.get()
    B_2n=B_2.get()
    B_3n=B_3.get()
    B_4n=B_4.get()
    Bcrop=(B_1n,B_2n,B_3n,B_4n)

    C_1n=C_1.get()
    C_2n=C_2.get()
    C_3n=C_3.get()
    C_4n=C_4.get()
    Ccrop=(C_1n,C_2n,C_3n,C_4n)

    D_1n=D_1.get()
    D_2n=D_2.get()
    D_3n=D_3.get()
    D_4n=D_4.get()
    Dcrop=(D_1n,D_2n,D_3n,D_4n)

    #ログ作成
    wb = openpyxl.load_workbook("system_log.xlsx")
    ws = wb['log']
    basename = os.path.basename(foldar_name)
    maxrow = ws.max_row
    if way ==0:
        wayname= "IF"
    elif way ==1:
        wayname= "GM"

    if Opening ==0:
        Openingname= "ON"
    elif Opening ==1:
        Openingname= "OFF"

    sheet = wb.active

    for x in ["A","B","C","D"]:
        excel_name = basename + "_"+x+"_"+ time
        maxrow = maxrow+1
        sheet.cell(row=maxrow, column=1).value = excel_name
        sheet.cell(row=maxrow, column=2).value = basename
        sheet.cell(row=maxrow, column=3).value = logtime
        sheet.cell(row=maxrow, column=4).value = first_num
        sheet.cell(row=maxrow, column=5).value = last_num
        sheet.cell(row=maxrow, column=6).value = wayname
        sheet.cell(row=maxrow, column=7).value = Openingname
        sheet.cell(row=maxrow, column=8).value = th



    wb.save('system_log.xlsx')


    if way == 0:
        if Opening == 0:
            mkdiffimg_def.mkdiffimgfiles_0_0(foldar_name,first_num,last_num,th,time)

        elif Opening == 1:
            mkdiffimg_def.mkdiffimgfiles_0_1(foldar_name,first_num,last_num,th,time)

    elif way == 1:
        if Opening == 0:
            mkdiffimg_def.mkdiffimgfiles_1_0(foldar_name,first_num,last_num,time)

        elif Opening == 1:
            mkdiffimg_def.mkdiffimgfiles_1_1(foldar_name,first_num,last_num,time)

    # メッセージボックス(コメントアウト)
    #messagebox.showinfo('差分作成完了', '差分作成完了しました。')
    print("Making difference images completed!")

    cropimage_def.cropimage_files(foldar_name,first_num,last_num,Acrop,Bcrop,Ccrop,Dcrop,time)
    # メッセージボックス(コメントアウト)
    #messagebox.showinfo('トリミング完了', 'トリミング完了しました。')
    print("Triming images completed!")

    diff_count_def.diffcountfiles(foldar_name,first_num,last_num,time)
    # メッセージボックス
    print("Counting pixels completed!")
    messagebox.showinfo('Processing', 'All Completed!!')

# メインウィンドウ
main_win = tkinter.Tk()
main_win.title("Quantitative activity system based on time-lapse images")
main_win.geometry("870x320")

# メインフレーム
main_frm = ttk.Frame(main_win)
main_frm.grid(column=0, row=0, sticky=tkinter.NSEW, padx=5, pady=10)

# パラメータ
folder_path = tkinter.StringVar()
excel_name = tkinter.StringVar()
first=tkinter.IntVar(value=1)
last=tkinter.IntVar(value=2)
threshold=tkinter.IntVar(value=40)

#ラジオボタン
radioValue = tkinter.IntVar()
Opening_value = tkinter.IntVar()


A_1=tkinter.IntVar()
A_2=tkinter.IntVar()
A_3=tkinter.IntVar(value=2000)
A_4=tkinter.IntVar(value=1000)

B_1=tkinter.IntVar()
B_2=tkinter.IntVar()
B_3=tkinter.IntVar(value=2000)
B_4=tkinter.IntVar(value=1000)

C_1=tkinter.IntVar()
C_2=tkinter.IntVar()
C_3=tkinter.IntVar(value=2000)
C_4=tkinter.IntVar(value=1000)

D_1=tkinter.IntVar()
D_2=tkinter.IntVar()
D_3=tkinter.IntVar(value=2000)
D_4=tkinter.IntVar(value=1000)

# ウィジェット（フォルダ名）
folder_label = ttk.Label(main_frm, text="Directory path")
folder_box = ttk.Entry(main_frm, textvariable=folder_path)
folder_btn = ttk.Button(main_frm, text="Reference",command= ask_folder)

#変数
first_num_label=ttk.Label(main_frm, text="First_num")
first_num_box=ttk.Entry(main_frm, textvariable=first)
last_num_label=ttk.Label(main_frm, text="Last_num")
last_num_box=ttk.Entry(main_frm, textvariable=last)
threshold_label=ttk.Label(main_frm, text="Threshold")
threshold_box=ttk.Entry(main_frm, textvariable=threshold)
excel_name_label=ttk.Label(main_frm, text="File_name")
excel_name_box=ttk.Entry(main_frm, textvariable=excel_name)

#画像作成方法選択
method_label = ttk.Label(main_frm, text="Method")
rdioOne = ttk.Radiobutton(main_frm, text='Interframe difference',variable=radioValue, value=0)
rdioTwo = ttk.Radiobutton(main_frm, text='Gaussian mixture',variable=radioValue, value=1)

#オープニング処理の有無
Opening_label = ttk.Label(main_frm, text="Opening")
Opening_On = ttk.Radiobutton(main_frm, text='On',variable=Opening_value, value=0)
Opening_Off = ttk.Radiobutton(main_frm, text='Off',variable=Opening_value, value=1)

Acrop_label=ttk.Label(main_frm, text="A")
Acrop_num_box1=ttk.Entry(main_frm, textvariable=A_1)
Acrop_num_box2=ttk.Entry(main_frm, textvariable=A_2)
Acrop_num_box3=ttk.Entry(main_frm, textvariable=A_3)
Acrop_num_box4=ttk.Entry(main_frm, textvariable=A_4)

Bcrop_label=ttk.Label(main_frm, text="B")
Bcrop_num_box1=ttk.Entry(main_frm, textvariable=B_1)
Bcrop_num_box2=ttk.Entry(main_frm, textvariable=B_2)
Bcrop_num_box3=ttk.Entry(main_frm, textvariable=B_3)
Bcrop_num_box4=ttk.Entry(main_frm, textvariable=B_4)

Ccrop_label=ttk.Label(main_frm, text="C")
Ccrop_num_box1=ttk.Entry(main_frm, textvariable=C_1)
Ccrop_num_box2=ttk.Entry(main_frm, textvariable=C_2)
Ccrop_num_box3=ttk.Entry(main_frm, textvariable=C_3)
Ccrop_num_box4=ttk.Entry(main_frm, textvariable=C_4)

Dcrop_label=ttk.Label(main_frm, text="D")
Dcrop_num_box1=ttk.Entry(main_frm, textvariable=D_1)
Dcrop_num_box2=ttk.Entry(main_frm, textvariable=D_2)
Dcrop_num_box3=ttk.Entry(main_frm, textvariable=D_3)
Dcrop_num_box4=ttk.Entry(main_frm, textvariable=D_4)
#値を取得

# ウィジェット（実行ボタン）
app_btn = ttk.Button(main_frm, text="START" ,command=app)

# ウィジェットの配置
folder_label.grid(column=0, row=0)
folder_box.grid(column=1, row=0,columnspan=2,padx=5,sticky=tkinter.W+tkinter.E )
folder_btn.grid(column=3, row=0)

method_label.grid(column=0, row=1)
rdioOne.grid(column=1, row=1,sticky=tkinter.W)
rdioTwo.grid(column=2, row=1,sticky=tkinter.W)

Opening_label.grid(column=0, row=2)
Opening_On.grid(column=1, row=2,sticky=tkinter.W)
Opening_Off.grid(column=2, row=2,sticky=tkinter.W)

threshold_label.grid(column=0, row=3)
threshold_box.grid(column=1, row=3,sticky=tkinter.W)

first_num_label.grid(column=0, row=4)
first_num_box.grid(column=1, row=4,sticky=tkinter.W)
last_num_label.grid(column=0, row=5)
last_num_box.grid(column=1, row=5,sticky=tkinter.W)

Acrop_label.grid(column=0, row=6)
Acrop_num_box1.grid(column=1, row=6,sticky=tkinter.W)
Acrop_num_box2.grid(column=2, row=6,sticky=tkinter.W)
Acrop_num_box3.grid(column=3, row=6,sticky=tkinter.W)
Acrop_num_box4.grid(column=4, row=6,sticky=tkinter.W)

Bcrop_label.grid(column=0, row=7)
Bcrop_num_box1.grid(column=1, row=7,sticky=tkinter.W)
Bcrop_num_box2.grid(column=2, row=7,sticky=tkinter.W)
Bcrop_num_box3.grid(column=3, row=7,sticky=tkinter.W)
Bcrop_num_box4.grid(column=4, row=7,sticky=tkinter.W)

Ccrop_label.grid(column=0, row=8)
Ccrop_num_box1.grid(column=1, row=8,sticky=tkinter.W)
Ccrop_num_box2.grid(column=2, row=8,sticky=tkinter.W)
Ccrop_num_box3.grid(column=3, row=8,sticky=tkinter.W)
Ccrop_num_box4.grid(column=4, row=8,sticky=tkinter.W)

Dcrop_label.grid(column=0, row=9)
Dcrop_num_box1.grid(column=1, row=9,sticky=tkinter.W)
Dcrop_num_box2.grid(column=2, row=9,sticky=tkinter.W)
Dcrop_num_box3.grid(column=3, row=9,sticky=tkinter.W)
Dcrop_num_box4.grid(column=4, row=9,sticky=tkinter.W)


app_btn.grid(column=1, row=10)

# 配置設定
main_win.columnconfigure(0, weight=1)
main_win.rowconfigure(0, weight=1)
main_frm.columnconfigure(1, weight=1)

main_win.mainloop()
